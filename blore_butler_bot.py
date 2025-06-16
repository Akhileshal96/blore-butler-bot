import telebot
import json
import openpyxl
import os

# =====================
# BOT CONFIGURATION
# =====================

# Your Bot Token
TOKEN = '7817780868:AAEd39YD3hDr1JAsQTCmeN9hgrMAtAHnKe4'
bot = telebot.TeleBot(TOKEN)

# Group ID (only group members can register)
GROUP_ID = -1001518197115

# Admins file
ADMINS_FILE = "admins.json"

# Create admins file if not exist
if not os.path.exists(ADMINS_FILE):
    with open(ADMINS_FILE, "w") as f:
        json.dump(["728623146"], f)

# Excel file for storing registrations
EXCEL_FILE = "registrations.xlsx"
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Telegram ID", "Telegram Name", "Full Name", "Phone Number"])
    wb.save(EXCEL_FILE)


# =====================
# HELPER FUNCTIONS
# =====================

# Load admins from file
def load_admins():
    with open(ADMINS_FILE, "r") as f:
        return json.load(f)

# Save admins to file
def save_admins(admins):
    with open(ADMINS_FILE, "w") as f:
        json.dump(admins, f)

# Check if user is group member
def is_group_member(user_id):
    try:
        member = bot.get_chat_member(GROUP_ID, user_id)
        return member.status in ['member', 'administrator', 'creator']
    except:
        return False


# =====================
# BOT HANDLERS
# =====================

# Start Command
@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.from_user.id

    if not is_group_member(user_id):
        bot.reply_to(message, "🚫 Sorry, you must be a member of Bangalore Meetups group to register.")
        return

    msg = bot.reply_to(message, "✅ Welcome to Bangalore Meetups Registration!\n\nPlease enter your full name:")
    bot.register_next_step_handler(msg, process_name)


# Process Full Name
def process_name(message):
    full_name = message.text.strip()
    user_id = message.from_user.id

    # Store full name temporarily
    bot.user_step = {}
    bot.user_step[user_id] = {'full_name': full_name}

    msg = bot.reply_to(message, "Please enter your 10-digit phone number:")
    bot.register_next_step_handler(msg, process_phone)


# Process Phone Number
def process_phone(message):
    phone = message.text.strip()
    user_id = message.from_user.id
    username = message.from_user.username or ""
    name = message.from_user.first_name or ""

    if not phone.isdigit() or len(phone) != 10:
        msg = bot.reply_to(message, "❌ Invalid phone number. Please enter a valid 10-digit phone number:")
        bot.register_next_step_handler(msg, process_phone)
        return

    full_name = bot.user_step[user_id]['full_name']

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([user_id, f"{name} (@{username})", full_name, phone])
    wb.save(EXCEL_FILE)

    bot.reply_to(message, "✅ You have been successfully registered for the Bangalore Meetups guest list!")


# Admin: Reset registrations
@bot.message_handler(commands=['reset'])
def reset(message):
    user_id = str(message.from_user.id)
    admins = load_admins()

    if user_id not in admins:
        bot.reply_to(message, "🚫 Unauthorized: Only admins can reset the list.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Telegram ID", "Telegram Name", "Full Name", "Phone Number"])
    wb.save(EXCEL_FILE)

    bot.reply_to(message, "✅ Registration list has been reset.")


# Admin: Add new admin
@bot.message_handler(commands=['addadmin'])
def addadmin(message):
    user_id = str(message.from_user.id)
    admins = load_admins()

    if user_id not in admins:
        bot.reply_to(message, "🚫 Unauthorized: Only admins can add new admins.")
        return

    try:
        new_admin_id = message.text.split(" ")[1]
        if new_admin_id not in admins:
            admins.append(new_admin_id)
            save_admins(admins)
            bot.reply_to(message, f"✅ Admin {new_admin_id} added successfully.")
        else:
            bot.reply_to(message, "ℹ️ User is already an admin.")
    except:
        bot.reply_to(message, "Usage: /addadmin <telegram_user_id>")


# Admin: Download registration data
@bot.message_handler(commands=['download'])
def download(message):
    user_id = str(message.from_user.id)
    admins = load_admins()

    if user_id not in admins:
        bot.reply_to(message, "🚫 Unauthorized: Only admins can download data.")
        return

    with open(EXCEL_FILE, "rb") as f:
        bot.send_document(message.chat.id, f)


# Fallback for unknown commands
@bot.message_handler(func=lambda m: True)
def fallback(message):
    bot.reply_to(message, "ℹ️ Use /start to begin registration.")


# Start bot polling
bot.infinity_polling()
